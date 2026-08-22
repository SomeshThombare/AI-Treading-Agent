"""
trades/reports/excel_generator.py

Generates Excel (.xlsx) portfolio report with multiple sheets:
  Sheet 1: Summary       — key statistics
  Sheet 2: Closed Trades — all closed trades with formulas
  Sheet 3: Open Trades   — current positions
  Sheet 4: PnL by Symbol — grouped performance
  Sheet 5: Bot Activity  — bot logs (if applicable)

Uses openpyxl library — produces editable spreadsheet
that opens in Excel, Google Sheets, or LibreOffice.
"""

import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

logger = logging.getLogger(__name__)


# ── Color palette ──────────────────────────────────────
COLOR_HEADER_BG  = '1F3864'
COLOR_HEADER_FG  = 'FFFFFF'
COLOR_ROW_ALT    = 'F2F2F2'
COLOR_GREEN      = '00B855'
COLOR_RED        = 'DC2626'
COLOR_YELLOW     = 'CA8A04'
COLOR_BLUE       = '0EA5E9'
COLOR_BORDER     = 'CBD5E1'
COLOR_TITLE_BG   = '0F1923'

# ── Styles ─────────────────────────────────────────────
THIN_BORDER = Border(
    left   = Side(style='thin', color=COLOR_BORDER),
    right  = Side(style='thin', color=COLOR_BORDER),
    top    = Side(style='thin', color=COLOR_BORDER),
    bottom = Side(style='thin', color=COLOR_BORDER),
)


# ─────────────────────────────────────────────────────
#  Main Generator Function
# ─────────────────────────────────────────────────────

def generate_excel_report(report_data):
    """
    Generate Excel report with multiple sheets.

    Args:
        report_data: dict from collect_report_data()

    Returns:
        BytesIO buffer with .xlsx bytes
    """
    logger.info("[EXCEL] Generating multi-sheet Excel report")

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Build all sheets
    _build_summary_sheet(    wb, report_data)
    _build_closed_trades_sheet(wb, report_data)
    _build_open_trades_sheet(  wb, report_data)
    _build_pnl_by_symbol_sheet(wb, report_data)

    # Bot sheet only if configured
    if report_data['bot']['configured']:
        _build_bot_sheet(wb, report_data)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(f"[EXCEL] Generated successfully — {len(buffer.getvalue())} bytes")
    return buffer


# ─────────────────────────────────────────────────────
#  Sheet 1: Summary
# ─────────────────────────────────────────────────────

def _build_summary_sheet(wb, data):
    """Top-level summary statistics sheet."""
    ws = wb.create_sheet('Summary', 0)
    ws.sheet_view.showGridLines = False

    # ── Title section ──
    ws.merge_cells('A1:D1')
    ws['A1'] = '⬡ AI TRADE — Portfolio Report'
    ws['A1'].font = Font(name='Calibri', size=22, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=COLOR_TITLE_BG)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # ── User and date range ──
    ws.merge_cells('A2:D2')
    user      = data['user']
    start     = data['start_date'].strftime('%b %d, %Y')
    end       = data['end_date'].strftime('%b %d, %Y')
    ws['A2']  = f'{user.username}   ·   {start} — {end}'
    ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='64748B')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25

    # ── Empty row ──
    row = 4

    # ── Section: Trade Statistics ──
    s = data['summary']
    ws.cell(row=row, column=1, value='TRADE STATISTICS').font = \
        Font(bold=True, size=11, color='FFFFFF')
    ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22
    row += 1

    metrics = [
        ('Total Trades',        s['total_trades']),
        ('Closed Trades',       s['closed_trades']),
        ('Open Trades',         s['open_trades']),
        ('Take Profit Wins',    s['wins']),
        ('Stop Loss Losses',    s['losses']),
        ('Manual Closes',       s['manual_closes']),
        ('Win Rate',            f"{s['win_rate']}%"),
        ('Total PnL',           f"${'+' if s['total_pnl'] >= 0 else ''}{s['total_pnl']}"),
        ('Average PnL',         f"${data['avg_pnl']}"),
    ]

    for label, value in metrics:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1)
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=str(value)).font = \
            Font(size=10, name='Consolas')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=2).border = THIN_BORDER

        # Alternating row color
        if row % 2 == 1:
            for col in (1, 2):
                ws.cell(row=row, column=col).fill = \
                    PatternFill('solid', fgColor=COLOR_ROW_ALT)
        row += 1

    row += 1

    # ── Section: Best / Worst trades ──
    ws.cell(row=row, column=1, value='BEST AND WORST TRADES').font = \
        Font(bold=True, size=11, color='FFFFFF')
    ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22
    row += 1

    if data['best_trade']:
        ws.cell(row=row, column=1, value='Best Trade').font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1)
        ws.cell(row=row, column=2,
                value=f"{data['best_trade']['symbol']}  "
                      f"(+{data['best_trade']['pnl']:.2f}%)")
        ws.cell(row=row, column=2).font = Font(bold=True, color=COLOR_GREEN, size=10)
        for c in (1, 2): ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    if data['worst_trade']:
        ws.cell(row=row, column=1, value='Worst Trade').font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1)
        ws.cell(row=row, column=2,
                value=f"{data['worst_trade']['symbol']}  "
                      f"({data['worst_trade']['pnl']:.2f}%)")
        ws.cell(row=row, column=2).font = Font(bold=True, color=COLOR_RED, size=10)
        for c in (1, 2): ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1

    # ── Section: Generated info ──
    gen_at = data['generated_at'].strftime('%B %d, %Y · %H:%M')
    ws.cell(row=row, column=1,
            value=f'Generated on {gen_at}').font = \
        Font(italic=True, size=9, color='64748B')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22


# ─────────────────────────────────────────────────────
#  Sheet 2: Closed Trades
# ─────────────────────────────────────────────────────

def _build_closed_trades_sheet(wb, data):
    """All closed trades with full details."""
    ws = wb.create_sheet('Closed Trades')

    headers = [
        '#', 'Symbol', 'Market', 'Entry Price', 'Exit Price',
        'TP %', 'SL %', 'Status', 'PnL %',
        'Opened', 'Closed', 'Duration (hours)'
    ]

    # ── Header row ──
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # ── Data rows ──
    row = 2
    for i, t in enumerate(data['closed_trades'], 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=t['symbol'])
        ws.cell(row=row, column=3, value=t['market'])
        ws.cell(row=row, column=4, value=t['entry_price'])
        ws.cell(row=row, column=5,
                value=t['current_price'] if t['current_price'] else '—')
        ws.cell(row=row, column=6, value=t['tp_percent'])
        ws.cell(row=row, column=7, value=t['sl_percent'])

        # Status with color
        status_text = t['status_display'] if 'status_display' in t else t['status']
        status_cell = ws.cell(row=row, column=8, value=status_text)
        if t['status'] == 'CLOSED_TP':
            status_cell.font = Font(bold=True, color=COLOR_GREEN, size=10)
        elif t['status'] == 'CLOSED_SL':
            status_cell.font = Font(bold=True, color=COLOR_RED, size=10)
        else:
            status_cell.font = Font(bold=True, color=COLOR_YELLOW, size=10)

        # PnL with color
        if t['pnl'] is not None:
            pnl_cell = ws.cell(row=row, column=9, value=t['pnl'])
            pnl_color = COLOR_GREEN if t['pnl'] >= 0 else COLOR_RED
            pnl_cell.font = Font(bold=True, color=pnl_color, size=10)
            # pnl_cell.number_format = '+0.00"%";-0.00"%";0.00"%"'
            pnl_cell.number_format = '"$"+0.00;"$"-0.00;"$"0.00'
        else:
            ws.cell(row=row, column=9, value='—')

        # Dates
        ws.cell(row=row, column=10,
                value=t['created_at'].strftime('%Y-%m-%d %H:%M') if t['created_at'] else '—')
        ws.cell(row=row, column=11,
                value=t['closed_at'].strftime('%Y-%m-%d %H:%M') if t['closed_at'] else '—')
        ws.cell(row=row, column=12,
                value=t['duration'] if t['duration'] is not None else '—')

        # Format prices as currency
        for col in (4, 5):
            if isinstance(ws.cell(row=row, column=col).value, (int, float)):
                ws.cell(row=row, column=col).number_format = '$#,##0.0000'

        # Format percents
        for col in (6, 7):
            ws.cell(row=row, column=col).number_format = '0.00"%"'

        # Alternating row background
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = \
                    PatternFill('solid', fgColor=COLOR_ROW_ALT)

        # Borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = \
                Alignment(horizontal='center', vertical='center')

        row += 1

    # ── Summary row at bottom (formulas) ──
    if data['closed_trades']:
        row += 1
        ws.cell(row=row, column=1, value='TOTAL').font = \
            Font(bold=True, size=10, color='FFFFFF')
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)

        # Sum formula for PnL column
        last_data_row = row - 2
        formula = f"=SUM(I2:I{last_data_row})"
        cell = ws.cell(row=row, column=9, value=formula)
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
        cell.number_format = '+0.00"%";-0.00"%";0.00"%"'

        # Fill rest with header color
        for col in range(2, 9):
            ws.cell(row=row, column=col).fill = \
                PatternFill('solid', fgColor=COLOR_HEADER_BG)

    # ── Column widths ──
    widths = [5, 12, 10, 14, 14, 9, 9, 14, 10, 18, 18, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze top row ──
    ws.freeze_panes = 'A2'

    # ── Add filter ──
    if data['closed_trades']:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row}"


# ─────────────────────────────────────────────────────
#  Sheet 3: Open Trades
# ─────────────────────────────────────────────────────

def _build_open_trades_sheet(wb, data):
    """Currently open positions."""
    ws = wb.create_sheet('Open Trades')

    headers = [
        '#', 'Symbol', 'Market', 'Entry Price', 'Current Price',
        'Take Profit', 'Stop Loss', 'TP %', 'SL %', 'Opened At'
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    row = 2
    for i, t in enumerate(data['open_trades'], 1):
        ws.cell(row=row, column=1,  value=i)
        ws.cell(row=row, column=2,  value=t['symbol'])
        ws.cell(row=row, column=3,  value=t['market'])
        ws.cell(row=row, column=4,  value=t['entry_price'])
        ws.cell(row=row, column=5,
                value=t['current_price'] if t['current_price'] else '—')
        ws.cell(row=row, column=6,  value=t['take_profit'])
        ws.cell(row=row, column=7,  value=t['stop_loss'])
        ws.cell(row=row, column=8,  value=t['tp_percent'])
        ws.cell(row=row, column=9,  value=t['sl_percent'])
        ws.cell(row=row, column=10,
                value=t['created_at'].strftime('%Y-%m-%d %H:%M'))

        # Format prices
        for col in (4, 5, 6, 7):
            if isinstance(ws.cell(row=row, column=col).value, (int, float)):
                ws.cell(row=row, column=col).number_format = '$#,##0.0000'

        # Format percents
        for col in (8, 9):
            ws.cell(row=row, column=col).number_format = '0.00"%"'

        # Color TP and SL columns
        ws.cell(row=row, column=6).font = Font(color=COLOR_GREEN, bold=True, size=10)
        ws.cell(row=row, column=7).font = Font(color=COLOR_RED,   bold=True, size=10)
        ws.cell(row=row, column=8).font = Font(color=COLOR_GREEN, size=10)
        ws.cell(row=row, column=9).font = Font(color=COLOR_RED,   size=10)

        # Alternating background
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = \
                    PatternFill('solid', fgColor=COLOR_ROW_ALT)

        # Borders + alignment
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = \
                Alignment(horizontal='center', vertical='center')

        row += 1

    if not data['open_trades']:
        ws.merge_cells('A2:J2')
        ws['A2'] = 'No open trades currently.'
        ws['A2'].font = Font(italic=True, size=10, color='64748B')
        ws['A2'].alignment = Alignment(horizontal='center')

    # Column widths
    widths = [5, 12, 10, 14, 14, 14, 14, 9, 9, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────────────
#  Sheet 4: PnL by Symbol
# ─────────────────────────────────────────────────────

def _build_pnl_by_symbol_sheet(wb, data):
    """Performance grouped by symbol with chart."""
    ws = wb.create_sheet('PnL by Symbol')

    headers = ['Symbol', 'Trade Count', 'Total PnL %', 'Avg PnL %']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
        cell.fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    row = 2
    for s in data['pnl_by_symbol']:
        ws.cell(row=row, column=1, value=s['symbol']).font = \
            Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=s['count'])
        ws.cell(row=row, column=3, value=s['pnl'])
        ws.cell(row=row, column=4,
                value=round(s['pnl'] / s['count'], 2) if s['count'] > 0 else 0)

        # Color PnL
        pnl_cell = ws.cell(row=row, column=3)
        avg_cell = ws.cell(row=row, column=4)
        color = COLOR_GREEN if s['pnl'] >= 0 else COLOR_RED
        pnl_cell.font = Font(bold=True, color=color, size=10)
        avg_cell.font = Font(color=color, size=10)
        # pnl_cell.number_format = '+0.00"%";-0.00"%";0.00"%"'
        pnl_cell.number_format = '"$"+0.00;"$"-0.00;"$"0.00'
        avg_cell.number_format = '+0.00"%";-0.00"%";0.00"%"'

        # Borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = \
                Alignment(horizontal='center', vertical='center')

        # Alternating
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                if col != 3 and col != 4:  # Don't override PnL color
                    ws.cell(row=row, column=col).fill = \
                        PatternFill('solid', fgColor=COLOR_ROW_ALT)


        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    # ── Add bar chart ──
    if data['pnl_by_symbol']:
        chart = BarChart()
        chart.type     = 'bar'
        chart.style    = 10
        chart.title    = 'PnL by Symbol'
        chart.y_axis.title = 'Symbol'
        chart.x_axis.title = 'PnL (%)'
        chart.height   = 10
        chart.width    = 16

        data_ref = Reference(ws, min_col=3, min_row=1, max_row=row-1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, 'F2')

    ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────────────
#  Sheet 5: Bot Activity
# ─────────────────────────────────────────────────────

def _build_bot_sheet(wb, data):
    """Bot statistics and activity log."""
    ws = wb.create_sheet('Bot Activity')
    ws.sheet_view.showGridLines = False

    bot = data['bot']

    # ── Title ──
    ws.merge_cells('A1:D1')
    ws['A1'] = '🤖 Auto Trading Bot'
    ws['A1'].font = Font(bold=True, size=18, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=COLOR_TITLE_BG)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # ── Bot stats section ──
    row = 3
    ws.cell(row=row, column=1, value='BOT STATISTICS').font = \
        Font(bold=True, size=11, color='FFFFFF')
    ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1

    status_text = 'ACTIVE ✓' if bot['is_active'] else 'INACTIVE'
    status_color = COLOR_GREEN if bot['is_active'] else '64748B'

    bot_stats = [
        ('Status',           status_text, status_color),
        ('Total Bot Trades', str(bot['total_trades']), '000000'),
        ('Wins',             str(bot['wins']),         COLOR_GREEN),
        ('Losses',           str(bot['losses']),       COLOR_RED),
        ('Win Rate',         f"{bot['win_rate']}%",
            COLOR_GREEN if bot['win_rate'] >= 50 else COLOR_RED),
        ('Max Open Trades',  str(bot['max_open_trades']),  '000000'),
        ('Crypto Threshold', f"{bot['min_conf_crypto']}%", '000000'),
        ('Forex Threshold',  f"{bot['min_conf_forex']}%",  '000000'),
        ('Gold Threshold',   f"{bot['min_conf_gold']}%",   '000000'),
        ('Selected Symbols', ', '.join(bot['selected_symbols'][:5]) or 'None', '000000'),
    ]

    for label, value, color in bot_stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1)
        ws.cell(row=row, column=1).border = THIN_BORDER

        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = Font(bold=True, color=color, size=10, name='Consolas')
        val_cell.alignment = Alignment(horizontal='left', indent=1)
        val_cell.border = THIN_BORDER

        if row % 2 == 1:
            for col in (1, 2):
                ws.cell(row=row, column=col).fill = \
                    PatternFill('solid', fgColor=COLOR_ROW_ALT)
        row += 1

    row += 2

    # ── Activity log section ──
    if bot['recent_logs']:
        ws.cell(row=row, column=1, value='RECENT ACTIVITY LOG').font = \
            Font(bold=True, size=11, color='FFFFFF')
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22
        row += 1

        log_headers = ['Timestamp', 'Action', 'Symbol', 'Message']
        for col, h in enumerate(log_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
            cell.fill = PatternFill('solid', fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        row += 1

        for log in bot['recent_logs']:
            ws.cell(row=row, column=1,
                    value=log['timestamp'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=2, value=log['action'])
            ws.cell(row=row, column=3, value=log['symbol'] or '—')
            ws.cell(row=row, column=4, value=log['message'])

            # Color action badge
            action_cell = ws.cell(row=row, column=2)
            action_cell.font = Font(bold=True, size=10)
            if log['action'] == 'OPEN':
                action_cell.font = Font(bold=True, color=COLOR_GREEN, size=10)
            elif log['action'] == 'CLOSE':
                action_cell.font = Font(bold=True, color=COLOR_YELLOW, size=10)
            elif log['action'] == 'ERROR':
                action_cell.font = Font(bold=True, color=COLOR_RED, size=10)
            elif log['action'] == 'SCAN':
                action_cell.font = Font(bold=True, color=COLOR_BLUE, size=10)

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = \
                    Alignment(horizontal='left', indent=1, vertical='center')

            if row % 2 == 0:
                for col in range(1, 5):
                    if col != 2:  # Don't override action color
                        ws.cell(row=row, column=col).fill = \
                            PatternFill('solid', fgColor=COLOR_ROW_ALT)
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 60